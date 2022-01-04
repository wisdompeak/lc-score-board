import json
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.drawing.image import Image
import argparse

###############################################

def convertToTitle(n):
  result = ''
  dist = ord('A') 
  while n > 0:
    y = (n-1) % 26
    n = (n-1) // 26
    s = chr(y+dist)
    result = ''.join((s, result))
  return result

mainArgumentsParser = argparse.ArgumentParser()
mainArgumentsParser.add_argument("endContest", help="EndContest id")
kwargs = mainArgumentsParser.parse_args()
endContest = int(kwargs.endContest)
        
startContest = endContest - 52 # as early as 118
ContestNumbers = endContest-startContest+1

###############################################

# Load company

from Data.Load_Company import LoadCompany

Company = LoadCompany()


###############################################

# Load Membership

from Data.Load_Membership import CalculateMembership

Membership = CalculateMembership()


###############################################

# import data
with open('../getRank/data.json', 'r') as f:
   	json_str = json.load(f)
   	data = json.loads(json_str)
   	
with open('../getRank/contests.json', 'r') as f:
   	json_str = json.load(f)
   	contests = json.loads(json_str) 

with open('../getRank/lc_profile_data.json', 'r') as f:
   	json_str = json.load(f)
   	profiles = json.loads(json_str)          	
   	
# import person   	
fi = open('../getRank/id.in', 'r')
id_list = [line.strip() for line in fi.readlines()]
fi.close()   	

# compute scores

table = []
for personID in id_list:  
  row = [personID] 
  for contestID in range(endContest,startContest-1,-1):
    contestID = str(contestID)    
    TotalPlayers = contests[contestID]
    if contestID not in data[personID]:
      row.append([-2,0,0])
    else:
      rank = data[personID][contestID][0]
      solved = data[personID][contestID][1]   
      if rank>0:   
      	score = round((1-rank/TotalPlayers)*80 + solved*5,1)
      else:
        score = 0
      row.append([rank,score,solved])

  print(personID)
  pool = [] # A pool for computing rolling score
  for x in row[1:4]:
    if x[0]!=-2: pool.append(x[1])
  if len(pool)<=2:
    RollingScore = round(sum(pool)/len(pool),1)
  else: 
    RollingScore = round((sum(pool)-min(pool))/2,1)

  if profiles[personID]["userContestRanking"]!=None:
    rating = profiles[personID]["userContestRanking"]["rating"]
  else:
    rating = -1
  newRow = [row[0], Membership[personID], round(rating), RollingScore]   # Reorder into [ID, days, rating, RollingScore,contestInfo]
  for x in row[1:]: newRow.append(x)
      
  table.append(newRow)  

table = sorted(table, key = lambda x:x[3], reverse = True)  

# for row in table:
#   print(row)
  	
############################

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'LeetCode Score Book'
colorChoice = ['EAEAEA','ffe6c2',colors.YELLOW,colors.GREEN,'19b457','ffb261']

sheet.column_dimensions['B'].width = 25.0
# sheet.column_dimensions['D'].width = 7.0
SIZE = 15

# YouXiu Logo
RowOffset = 1
row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column)+str(row+1) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/youxiu.html'
sheet[idx].value = "YouXiu-Logo"

# output header "Wisdompeak Cup"
row, column = RowOffset+1, 3
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+15)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet[idx1].value = 'Happy New Year! See 2021 Year-end Review!'
sheet[idx1].hyperlink = 'https://docs.google.com/presentation/d/1wIYIszpPuKEKTMAQXyb4lllhXApNWfa9J53lYK5gHmM/'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True, size=SIZE)

# output header "2020 year-end review"
row, column = RowOffset+2, 3
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+15)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet[idx1].value = 'Looking for more challenges? Try Wisdom Cup!'
sheet[idx1].hyperlink = 'https://wisdompeak.github.io/lc-score-board/cup.html'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True, size=SIZE)

# '''Kuai Shou Sponsorship'''
# # Kuaishou Logo
# RowOffset += 3
# row, column = RowOffset+1, 2
# idx = convertToTitle(column)+str(row) 
# idx2 = convertToTitle(column)+str(row+1) 
# sheet.merge_cells(idx+':'+idx2)
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
# sheet[idx].value = "Kuaishou-Big-Logo"

# # output header "Wisdompeak Cup"
# row, column = RowOffset+1, 3
# idx1 = convertToTitle(column)+str(row)
# idx2 = convertToTitle(column+20)+str(row)
# sheet.merge_cells(idx1+':'+idx2)
# sheet[idx1].value = 'Special thanks to Kuaishou Tech and MMU Team for sponsoring our club activities!'
# sheet[idx1].alignment = Alignment(horizontal='left')
# sheet[idx1].font = Font(bold=True, size=SIZE)

# # output header "MMU poster"
# row, column = RowOffset+2, 3
# idx1 = convertToTitle(column)+str(row)
# idx2 = convertToTitle(column+20)+str(row)
# sheet.merge_cells(idx1+':'+idx2)
# sheet[idx1].value = 'Click to check out their open positions in Beijing'
# sheet[idx1].hyperlink = 'https://wisdompeak.github.io/lc-score-board/Img/MMU.jpg'
# sheet[idx1].alignment = Alignment(horizontal='left')
# sheet[idx1].font = Font(bold=True, size=SIZE)

'''End of Kuai Shou Sponsorship'''

# output header "Contest Score Board"
RowOffset += 3
row, column = RowOffset+1, 2
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+15)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet.row_dimensions[row].height = 20.0
sheet[idx1].value = 'LeetCode Weekly Contest Score Board (the most recent 52 weeks)'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True, size=18)

RowOffset += 2
row, column = RowOffset+1, 2
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+3)+str(row)
sheet.merge_cells(idx1+':'+idx2)
# sheet.row_dimensions[row].height = 20.0
sheet[idx1].value = 'You may sort this table using links here:'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True)
sheet[idx1].font = Font(bold=True, size=12)

row, column = RowOffset+1, 6
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+2)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet[idx1].value = 'by @delphih'
sheet[idx1].hyperlink = 'http://rank.cruelcoding.com/'
sheet[idx1].alignment = Alignment(horizontal='center')
sheet[idx1].font = Font(bold=True, size=12)
row, column = RowOffset+1, 9
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+2)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet[idx1].value = 'by @rayms'
sheet[idx1].hyperlink = 'http://cruelcoding.com/'
sheet[idx1].alignment = Alignment(horizontal='center')
sheet[idx1].font = Font(bold=True, size=12)

RowOffset += 2

# output header "contest" 
row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Contest'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "contest" "Participants"
row, column = RowOffset+2, 2
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Participants'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "Days"
row, column = RowOffset+2, 3
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Days'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "Rating"
row, column = RowOffset+2, 4
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Rating'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "Score"
row, column = RowOffset+2, 5
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Score'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output contest ID / number of players
for k in range(endContest-startContest+1):
  row, column = RowOffset+1, 6+k*2
  idx1 = convertToTitle(column)+str(row)
  idx2 = convertToTitle(column+1)+str(row)  
  sheet.merge_cells(idx1+':'+idx2)   
  sheet[idx1].value = endContest-k 
  sheet[idx1].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx1].font = Font(bold=True, size=SIZE)
  sheet[idx1].fill = PatternFill("solid", fgColor='D9D9D9')
  
  row, column = RowOffset+2, 6+k*2
  idx1 = convertToTitle(column)+str(row)
  idx2 = convertToTitle(column+1)+str(row)  
  sheet.merge_cells(idx1+':'+idx2)   
  sheet[idx1].value = contests[str(endContest-k)] 
  sheet[idx1].font = Font(size=SIZE)
  sheet[idx1].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx1].fill = PatternFill("solid", fgColor='D9D9D9')


RowOffset += 4

for i in range(len(id_list)):

  row = RowOffset+i

  # 1st column: Rank
  row, column = RowOffset+i, 1
  idx = convertToTitle(column)+str(row) 
  sheet[idx].value = i+1
  sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx].font = Font(bold=True, size=SIZE)       
  if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')     
  name = table[i][0] 
  
  for j in range(len(table[i])):  # column index
     
    # 2nd Column: LC ID
    if j==0:   
      row, column = RowOffset+i, 2
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j]
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(bold=True, size=SIZE)       
      sheet[idx].hyperlink = 'http://leetcode.com/'+table[i][j]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA') 

    # 3rd column :  days
    elif j==1:   
      row, column = RowOffset+i, 3
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j]
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=12)
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')           

    # 4th column :  rating
    elif j==2:   
      row, column = RowOffset+i, 4
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j]
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA') 
      if (table[i][j]>=2100):
        sheet[idx].font = Font(size=12,color='A30000')
      else:
        sheet[idx].font = Font(size=12,color='0426A4')

    # 5th column :  avg score
    elif j==3:            
      row, column = RowOffset+i, 5
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j]
      sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[5])
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
      
    else:     # output rank and score                    
      row, column = RowOffset+i, 6+(j-4)*2      
      idx = convertToTitle(column)+str(row) 
      
      sheet[idx].value = table[i][j][0]
      for k in range(0,5):
        if k>0 and table[i][j][2]==k or k==0 and i%2==0:  
          sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[k])      
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
      
      row, column = RowOffset+i, 6+(j-4)*2+1
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j][1]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')      
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
              
  #### Company Logo    
  if name in Company:          
    row, column = RowOffset+i, 6+ContestNumbers*2
    idx = convertToTitle(column)+str(row)
    sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
    sheet[idx].value = Company[name]+"-Logo"


############################

RowOffset += len(id_list)

### Lee215

# # ID
# row, column = RowOffset+1, 2        
# idx = convertToTitle(column)+str(row) 
# sheet[idx].value = 'lee215'
# sheet[idx].hyperlink = 'https://leetcode.com/lee215'
# sheet[idx].font = Font(bold=True, size=SIZE) 
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
# sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')     

# # Days
# row, column = RowOffset+1, 3
# idx = convertToTitle(column)+str(row) 
# sheet[idx].value = Membership['lee215']
# sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
# sheet[idx].font = Font(size=12)

# # Rating
# row, column = RowOffset+1, 4
# idx = convertToTitle(column)+str(row) 
# sheet[idx].value = -1
# sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
# sheet[idx].font = Font(size=12)

# # Score
# row, column = RowOffset+1, 5        
# idx = convertToTitle(column)+str(row) 
# sheet[idx].value = 'YouXiu'
# sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[5])
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
# sheet[idx].font = Font(size=SIZE)

# # Contests Info
# row, column = RowOffset+1, 6
# idx = convertToTitle(column)+str(row) 
# idx2 = convertToTitle(column+ContestNumbers*2-1)+str(row) 
# sheet.merge_cells(idx+':'+idx2)
# sheet[idx].alignment = Alignment(horizontal='left',vertical='center')
# sheet[idx].value = 'Why always so YOUXIU?'
# sheet[idx].font = Font(bold=True, size=SIZE)
# sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/youxiu.html'
# sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')

# # Logo
# row, column = RowOffset+1, 6+ContestNumbers*2
# idx = convertToTitle(column)+str(row) 
# sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
# sheet[idx].value = Company["lee215"]+"-Logo"

############################
      
### legend
      
RowOffset += 2

row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = '-1:     absent/zero'
sheet[idx].font = Font(size=15)

row, column = RowOffset+2, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = '-2:     not joined'
sheet[idx].font = Font(size=15)

row, column = RowOffset+3, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = 'BG Color: # of problems solved.'
sheet[idx].font = Font(size=15)

RowOffset += 5
row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Full list of daily problems'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://bit.ly/2X0NW4e'

row, column = RowOffset+2, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Recommended resources'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/resources.html'

row, column = RowOffset+3, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'See where we are from'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://www.google.com/maps/d/viewer?mid=1c1t3qKsKxTTnDD_P2BsFsrL7p8l9Bou8'


row, column = RowOffset+4, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Cruel System Design Group'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/sd.html'


row, column = RowOffset+5, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'If you are interested in joining this group, please contact: guan.huifeng@gmail.com'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)

row, column = RowOffset+6, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Make sure you agree with our terms and regulations. Think twice before you apply.'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
# sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/rules.html'

row, column = RowOffset+7, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = '[Blacklisted Persons]'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/Blacklist/Darren_Zang.html'

############################

RowOffset += 9
row, column = RowOffset, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Graduated members and archived scores'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/graduates.html'

### curve figure

RowOffset += 2

row, column = RowOffset, 1
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+80)+str(row)
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'curve-figure'


############################
print("index.xlsx generated.")
wb.save('index.xlsx')    	

