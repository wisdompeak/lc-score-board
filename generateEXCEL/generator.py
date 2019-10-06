import json
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.drawing.image import Image



###############################################

def convertToTitle(n):
  result = ''
  dist = ord('A') 
  while n > 0:
    y = (n-1) % 26
    n = (n-1) // 26
    s = chr(y+dist)
    result = ''.join((s, result))
  return result
        
startContest = 118
endContest = 157
ContestNumbers = endContest-startContest+1

###############################################

# Load company

from Data.Load_Company import LoadCompany

Company = LoadCompany()


###############################################

# Load Membership

from Data.Load_Membership import CalculateMembership

Membership = CalculateMembership()


###############################################

# import data
with open('../getRank/data.json', 'r') as f:
   	json_str = json.load(f)
   	data = json.loads(json_str)
   	
with open('../getRank/contests.json', 'r') as f:
   	json_str = json.load(f)
   	contests = json.loads(json_str)     	
   	
# import person   	
fi = open('../getRank/id.in', 'r')
id_list = [line.strip() for line in fi.readlines()]
id_list.remove("lee215")
fi.close()   	

# compute scores

table = []
for personID in id_list:  
  row = [personID] 
  for contestID in range(endContest,startContest-1,-1):
    contestID = str(contestID)    
    TotalPlayers = contests[contestID]
    if contestID not in data[personID]:
      row.append([-2,0,0])
    else:
      rank = data[personID][contestID][0]
      solved = data[personID][contestID][1]   
      if rank>0:   
      	score = round((1-rank/TotalPlayers)*80 + solved*5,1)
      else:
        score = 0
      row.append([rank,score,solved])
   
  pool = [] # A pool for computing rolling score
  for x in row[1:4]:
    if x[0]!=-2: pool.append(x[1])
  if len(pool)<=2:
    RollingScore = round(sum(pool)/len(pool),1)
  else: 
    RollingScore = round((sum(pool)-min(pool))/2,1)
        
  newRow = [row[0],RollingScore]   # Reorder into [ID,RollingScore,contestInfo]
  for x in row[1:]: newRow.append(x)
      
  table.append(newRow)  

table = sorted(table, key = lambda x:x[1], reverse = True)  

for row in table:
  print(row)
  
############################ Graduated or laidoff members

table2 = []
for personID in data:

  if personID in id_list: continue
  if personID=="Ansonluo": continue
  if personID=="": continue
  
  row = [personID] 
  for contestID in range(endContest,startContest-1,-1):
    contestID = str(contestID)
    TotalPlayers = contests[contestID]
    if contestID not in data[personID]:
      row.append([-3,0,0])
    else:
      rank = data[personID][contestID][0]
      solved = data[personID][contestID][1]   
      if rank>0:   
      	score = round((1-rank/TotalPlayers)*80 + solved*5,1)
      else:
        score = 0
      row.append([rank,score,solved])
   
  RollingScore = 0       
  newRow = [row[0],RollingScore]   # Reorder into [ID,RollingScore,contestInfo]
  for x in row[1:]: newRow.append(x)
      
  table2.append(newRow)  

table2 = sorted(table2, key = lambda x:x[0])  

print("**********************************************************")
for row in table2:
  print(row)

	
############################

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'LeetCode Score Book'
colorChoice = ['EAEAEA','ffe6c2',colors.YELLOW,colors.GREEN,'19b457','ffb261']

sheet.column_dimensions['B'].width = 25.0
# sheet.column_dimensions['D'].width = 7.0
SIZE = 15

# output header "Wisdompeak Cup"
RowOffset = 1
row, column = RowOffset+1, 2
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+8)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet[idx1].value = 'Wanna more challenge? Try Wisdom Cup!'
sheet[idx1].hyperlink = 'https://wisdompeak.github.io/lc-score-board/cup.html'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True, size=SIZE)

# output header "Contest Score Board"
RowOffset += 2
row, column = RowOffset+1, 2
idx1 = convertToTitle(column)+str(row)
idx2 = convertToTitle(column+8)+str(row)
sheet.merge_cells(idx1+':'+idx2)
sheet.row_dimensions[row].height = 20.0
sheet[idx1].value = 'LeetCode Weekly Contest Score Board'
sheet[idx1].alignment = Alignment(horizontal='left')
sheet[idx1].font = Font(bold=True, size=18)


RowOffset += 2;

# output header "contest" 
row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Contest'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "contest" "Participants"
row, column = RowOffset+2, 2
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Participants'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "Score"
row, column = RowOffset+2, 3
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Days'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output header "Membership"
row, column = RowOffset+2, 4
idx = convertToTitle(column)+str(row)
sheet[idx].value = 'Score'
sheet[idx].alignment = Alignment(horizontal='center')
sheet[idx].font = Font(bold=True, size=SIZE)

# output contest ID / number of players
for k in range(endContest-startContest+1):
  row, column = RowOffset+1, 5+k*2
  idx1 = convertToTitle(column)+str(row)
  idx2 = convertToTitle(column+1)+str(row)  
  sheet.merge_cells(idx1+':'+idx2)   
  sheet[idx1].value = endContest-k 
  sheet[idx1].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx1].font = Font(bold=True, size=SIZE)
  sheet[idx1].fill = PatternFill("solid", fgColor='D9D9D9')
  
  row, column = RowOffset+2, 5+k*2
  idx1 = convertToTitle(column)+str(row)
  idx2 = convertToTitle(column+1)+str(row)  
  sheet.merge_cells(idx1+':'+idx2)   
  sheet[idx1].value = contests[str(endContest-k)] 
  sheet[idx1].font = Font(size=SIZE)
  sheet[idx1].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx1].fill = PatternFill("solid", fgColor='D9D9D9')


RowOffset += 4;

for i in range(len(id_list)):

  row = RowOffset+i

  # 1st column: Rank
  row, column = RowOffset+i, 1
  idx = convertToTitle(column)+str(row) 
  sheet[idx].value = i+1
  sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
  sheet[idx].font = Font(bold=True, size=SIZE)       
  if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')     
  name = table[i][0] 
  
  for j in range(len(table[i])):  # column index
     
    # 2nd Column: LC ID
    if j==0:   
      row, column = RowOffset+i, 2
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j]
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(bold=True, size=SIZE)       
      sheet[idx].hyperlink = 'http://leetcode.com/'+table[i][j]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')   
        
    # 3rd column :  avg score
    elif j==1:            
      row, column = RowOffset+i, 4
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][1]
      sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[5])
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
      
    else:     # output rank and score                    
      row, column = RowOffset+i, 5+(j-2)*2      
      idx = convertToTitle(column)+str(row) 
      
      sheet[idx].value = table[i][j][0]
      for k in range(0,5):
        if k>0 and table[i][j][2]==k or k==0 and i%2==0:  
          sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[k])      
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
      
      row, column = RowOffset+i, 5+(j-2)*2+1
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table[i][j][1]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')      
      sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
      sheet[idx].font = Font(size=SIZE)
  
  #### Membership
      
  row, column = RowOffset+i, 3
  idx = convertToTitle(column)+str(row)
  sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
  sheet[idx].value = Membership[name]
  if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')   
  sheet[idx].font = Font(size=12)
            
  #### Company Logo    
  if name in Company:          
    row, column = RowOffset+i, 5+ContestNumbers*2
    idx = convertToTitle(column)+str(row)
    sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
    sheet[idx].value = Company[name]+"-Logo"


############################

### Lee215

RowOffset += len(id_list)

row, column = RowOffset+1, 2        
idx = convertToTitle(column)+str(row) 
sheet[idx].value = 'lee215'
sheet[idx].hyperlink = 'https://leetcode.com/lee215'
sheet[idx].font = Font(bold=True, size=SIZE) 
sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')     

row, column = RowOffset+1, 3
idx = convertToTitle(column)+str(row) 
sheet[idx].value = Membership['lee215']
sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')
sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
sheet[idx].font = Font(size=12)

row, column = RowOffset+1, 4        
idx = convertToTitle(column)+str(row) 
sheet[idx].value = 'YouXiu'
sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[5])
sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
sheet[idx].font = Font(size=SIZE)

row, column = RowOffset+1, 5+ContestNumbers*2
idx = convertToTitle(column)+str(row) 
sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
sheet[idx].value = Company["lee215"]+"-Logo"

row, column = RowOffset+1, 5
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+ContestNumbers*2-1)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].alignment = Alignment(horizontal='center',vertical='center')
sheet[idx].value = 'Why always so YOUXIU?'
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/youxiu.html'
sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')


############################
      
### legend
      
RowOffset += 2

row, column = RowOffset+1, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = '-1:     absent/zero'
sheet[idx].font = Font(size=15)

row, column = RowOffset+2, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = '-2:     not joined'
sheet[idx].font = Font(size=15)

row, column = RowOffset+3, 2
idx = convertToTitle(column)+str(row) 
sheet[idx].value = '-3:     graduated/quited'
sheet[idx].font = Font(size=15)

row, column = RowOffset+5, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Full list of daily problems'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://bit.ly/2X0NW4e'

row, column = RowOffset+6, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Recommended resources'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/resources.html'

row, column = RowOffset+7, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'See where we are from'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://drive.google.com/open?id=1c1t3qKsKxTTnDD_P2BsFsrL7p8l9Bou8&usp=sharing'


row, column = RowOffset+8, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+5)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Cruel System Design Group'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/sd.html'


row, column = RowOffset+10, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'If you are interested in joining this group, please contact: guan.huifeng@gmail.com'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)

row, column = RowOffset+11, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+15)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Make sure you agree with our terms and regulations'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)
sheet[idx].hyperlink = 'https://wisdompeak.github.io/lc-score-board/rules.html'

############################

### curve figure

RowOffset += 12

row, column = RowOffset, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+60)+str(row)
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'curve-figure'


############################

RowOffset += 3
row, column = RowOffset, 2
idx = convertToTitle(column)+str(row) 
idx2 = convertToTitle(column+8)+str(row) 
sheet.merge_cells(idx+':'+idx2)
sheet[idx].value = 'Graduated or quited members in 2019'
sheet[idx].font = Font(size=15)
sheet[idx].font = Font(bold=True, size=SIZE)


RowOffset += 2

for i in range(len(table2)):

  # 1st column: Rank
  row, column = RowOffset+i, 1
  idx = convertToTitle(column)+str(row) 
  sheet[idx].value = "-"
  sheet[idx].alignment = Alignment(horizontal='center') 
  sheet[idx].font = Font(bold=True, size=SIZE)       
  if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')      
  name = table2[i][0]

  
  for j in range(len(table2[i])):  # column index
     
    # 2nd Column: LC ID
    if j==0:   
      row, column = RowOffset+i, 2
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table2[i][0]
      sheet[idx].alignment = Alignment(horizontal='center') 
      sheet[idx].font = Font(bold=True, size=SIZE)       
      sheet[idx].hyperlink = 'http://leetcode.com/'+table2[i][j]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')   
        
    # 3rd column :  avg score
    elif j==1:            
      row, column = RowOffset+i, 4
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = "-"
      sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[5])
      sheet[idx].alignment = Alignment(horizontal='center')
      sheet[idx].font = Font(size=SIZE)
      
    else:     # output rank and score                    
      row, column = RowOffset+i, 5+(j-2)*2      
      idx = convertToTitle(column)+str(row) 
      
      sheet[idx].value = table2[i][j][0]
      for k in range(0,5):
        if k>0 and table2[i][j][2]==k or k==0 and i%2==0:  
          sheet[idx].fill = PatternFill("solid", fgColor=colorChoice[k])      
      sheet[idx].alignment = Alignment(horizontal='center') 
      sheet[idx].font = Font(size=SIZE)
      
      row, column = RowOffset+i, 5+(j-2)*2+1
      idx = convertToTitle(column)+str(row) 
      sheet[idx].value = table2[i][j][1]
      if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')      
      sheet[idx].alignment = Alignment(horizontal='center') 
      sheet[idx].font = Font(size=SIZE)
      
      
  #### Membership
      
  row, column = RowOffset+i, 3
  idx = convertToTitle(column)+str(row)
  sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
  sheet[idx].value = Membership[name]
  if (i%2==0):
        sheet[idx].fill = PatternFill("solid", fgColor='EAEAEA')   
  sheet[idx].font = Font(size=12)
            
  #### Company Logo    
  if name in Company:          
    row, column = RowOffset+i, 5+ContestNumbers*2
    idx = convertToTitle(column)+str(row)
    sheet[idx].alignment = Alignment(horizontal='center',vertical='center') 
    sheet[idx].value = Company[name]+"-Logo"


############################

    
wb.save('index.xlsx')    	

